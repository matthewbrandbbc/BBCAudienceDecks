from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

from .config import (
    ALL_MARKETS_FOOTER,
    DATA_DIR,
    MIN_RESPONSES,
    PLATFORM_COLUMNS,
    PLATFORM_FOOTERS,
    REGION_FILE_HINTS,
)


class ValidationError(RuntimeError):
    """Raised when a source workbook violates the agreed contract."""


def clean_audience_name(value: str) -> str:
    value = re.sub(r"^\s*(?:Base:\s*)?", "", str(value or ""), flags=re.I)
    value = re.sub(r"^\s*AD\s*-\s*", "", value, flags=re.I)
    value = re.sub(r"\s*\(Audience Size\)\s*$", "", value, flags=re.I)
    value = value.strip(" .")
    if value.lower() in {"totals", "all internet users", "all internet users (audience"}:
        return "All Audiences"
    replacements = {
        "C-suites": "C-Suites",
        "C-Suites": "C-Suites",
        "Hnwis": "HNWIs",
        "Smes": "SMEs",
        "Business Decision Makers (Bdms)": "Business Decision Makers (BDMs)",
        "Fbdms - Finance Business Decision Makers": "FBDMs – Finance Business Decision Makers",
        "It Business Decision Makers (Itbdms)": "IT Business Decision Makers (ITBDMs)",
        "Int. Leisure Travellers": "International Leisure Travellers",
    }
    pretty = value.title()
    for source, target in replacements.items():
        if pretty.casefold() == source.casefold() or value.casefold() == source.casefold():
            return target
    return pretty


def normalize_for_match(value: str) -> str:
    value = re.sub(r"^\s*AD\s*-\s*", "", str(value or ""), flags=re.I)
    value = re.sub(r"\([^)]*\)", "", value)
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def format_compact(value: float | int) -> str:
    value = float(value)
    if abs(value) >= 1_000_000:
        return f"{round(value / 1_000_000):.0f}M"
    return f"{round(value / 1_000):.0f}K"


def format_percent(value: float | int) -> str:
    value = float(value)
    return f"{round(value * 100 if abs(value) <= 1 else value):.0f}"


def format_index(value: float | int) -> str:
    return f"{round(float(value)):.0f}"


def format_index_difference(value: float | int) -> str:
    result = round(float(value) - 100)
    return f"+{result}" if result > 0 else str(result)


def clean_competitor_name(value: str) -> str:
    text = str(value or "")
    text = re.sub(r"^\s*2(?:\.0)?\s+", "", text, flags=re.I)
    text = re.sub(r"\b(?:Cross Platform|Digital|TV)\b", "", text, flags=re.I)
    text = re.sub(r"\bEngagement\b", "", text, flags=re.I)
    return re.sub(r"\s+", " ", text).strip(" -")


def ordinal(value: int) -> str:
    if 10 <= value % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(value % 10, "th")
    return f"{value}{suffix}"


@dataclass(frozen=True)
class Competitor:
    name: str
    value: float


def sort_competitors(items: Iterable[Competitor]) -> list[Competitor]:
    # Stable, deterministic tie rule: descending value, then name alphabetically.
    return sorted(items, key=lambda item: (-item.value, item.name.casefold()))


class WorkbookRepository:
    def __init__(self, data_dir: Path = DATA_DIR):
        self.data_dir = Path(data_dir)
        self.paths = self._discover_paths()
        self._books = {}

    def _discover_paths(self) -> dict[str, Path]:
        found = {}
        files = list(self.data_dir.glob("*.xlsx"))
        for region, hint in REGION_FILE_HINTS.items():
            matches = [p for p in files if hint.casefold() in p.name.casefold()]
            if len(matches) != 1:
                raise ValidationError(
                    f"Expected exactly one workbook for {region!r} containing {hint!r}; found {len(matches)}."
                )
            found[region] = matches[0]
        return found

    def workbook(self, region: str):
        if region not in self.paths:
            raise ValidationError(f"Unknown region: {region}")
        if region not in self._books:
            # These are small, fixed-shape exports. Normal mode makes repeated
            # validated cell lookups dramatically faster than read-only mode.
            self._books[region] = load_workbook(self.paths[region], data_only=True, read_only=False)
        return self._books[region]

    def audience_map(self, region: str = "All Markets") -> dict[str, str]:
        wb = self.workbook(region)
        result = {}
        for ws in wb.worksheets:
            raw = str(ws["B7"].value or ws.title).replace("Base:", "").strip()
            cleaned = clean_audience_name(raw)
            if "all internet users" in raw.casefold():
                cleaned = "All Audiences"
            if cleaned in result:
                raise ValidationError(f"Duplicate cleaned audience name {cleaned!r} in {self.paths[region].name}.")
            result[cleaned] = ws.title
        return result

    def sheet(self, region: str, audience: str):
        mapping = self.audience_map(region)
        if audience not in mapping:
            raise ValidationError(f"Audience {audience!r} does not exist in {self.paths[region].name}.")
        return self.workbook(region)[mapping[audience]]

    def all_audience_sheet(self, region: str):
        return self.sheet(region, "All Audiences")

    def value(self, region: str, audience: str, cell: str):
        value = self.sheet(region, audience)[cell].value
        if value is None:
            raise ValidationError(f"Blank required cell {cell} in {self.paths[region].name} / {audience}.")
        return value

    def response_count(self, region: str, audience: str, platform: str, row: int) -> int:
        cell = f"{PLATFORM_COLUMNS[platform]['responses']}{row}"
        value = self.value(region, audience, cell)
        if not isinstance(value, (int, float)):
            raise ValidationError(f"Response count {cell} is not numeric.")
        return int(value)

    def is_supported(self, region: str, audience: str, platform: str, row: int) -> bool:
        return self.response_count(region, audience, platform, row) >= MIN_RESPONSES

    def total_response_count(self, region: str, audience: str, row: int) -> int:
        value = self.value(region, audience, f"E{row}")
        if not isinstance(value, (int, float)):
            raise ValidationError(f"Total response count E{row} is not numeric.")
        return int(value)

    def total_is_supported(self, region: str, audience: str, row: int) -> bool:
        return self.total_response_count(region, audience, row) >= MIN_RESPONSES

    def markets_footer(self, region: str) -> str:
        if region == "All Markets":
            return ALL_MARKETS_FOOTER
        raw = str(self.sheet(region, "All Audiences")["B8"].value or "")
        countries = re.sub(r"^\s*Countries:\s*", "", raw, flags=re.I).strip()
        if region == "Europe":
            countries = re.sub(r"(?:,\s*)?Russia(?:,\s*)?", ", ", countries, flags=re.I)
            countries = re.sub(r",\s*,", ",", countries).strip(" ,")
            return f"Markets: {countries} (Excl. Russia)"
        return f"Markets: {countries}"

    def waves(self, region: str, audience: str) -> str:
        raw = str(self.sheet(region, audience)["B9"].value or "")
        waves = re.sub(r"^\s*Waves:\s*", "", raw, flags=re.I).strip()
        return f"Source: GWI Core Survey — Waves: {waves}"

    def slide2_affinity(self, region: str, audience: str) -> float:
        ws = self.all_audience_sheet(region)
        selected_raw = str(self.sheet(region, audience)["B7"].value or "").replace("Base:", "").strip()
        target = "totals" if audience == "All Audiences" else normalize_for_match(selected_raw)
        matches = []
        for row in range(16, 33):
            label = str(ws.cell(row=row, column=3).value or "")
            normalized = label.casefold().strip() if audience == "All Audiences" else normalize_for_match(label)
            if normalized == target:
                matches.append(ws.cell(row=row, column=13).value)
        if len(matches) != 1:
            raise ValidationError(
                f"Slide 2 lookup expected one match for {audience!r} in {self.paths[region].name}; found {len(matches)}."
            )
        if not isinstance(matches[0], (int, float)):
            raise ValidationError(f"Slide 2 affinity for {audience!r} is not numeric.")
        return float(matches[0])

    def competitor_set(self, region: str, audience: str, start_row: int) -> list[Competitor]:
        ws = self.sheet(region, audience)
        items = []
        for row in range(start_row, start_row + 6):
            label = clean_competitor_name(ws[f"C{row}"].value)
            value = ws[f"D{row}"].value
            if not label or not isinstance(value, (int, float)):
                raise ValidationError(f"Invalid competitor data at C{row}:D{row} in {self.paths[region].name}.")
            items.append(Competitor(label, float(value)))
        sorted_items = sort_competitors(items)
        if sum(1 for item in sorted_items if item.name.casefold() == "bbc") != 1:
            raise ValidationError(f"Competitive set at rows {start_row}:{start_row+5} must contain BBC exactly once.")
        return sorted_items

    def supported_competitor_set(
        self, region: str, audience: str, start_row: int
    ) -> list[Competitor]:
        """Return only competitor results with at least 50 total responses."""
        ws = self.sheet(region, audience)
        items = []
        for row in range(start_row, start_row + 6):
            if not self.total_is_supported(region, audience, row):
                continue
            label = clean_competitor_name(ws[f"C{row}"].value)
            value = ws[f"D{row}"].value
            if label and isinstance(value, (int, float)):
                items.append(Competitor(label, float(value)))
        return sort_competitors(items)

    def validate_all(self) -> list[str]:
        errors = []
        baseline = None
        for region in REGION_FILE_HINTS:
            try:
                mapping = self.audience_map(region)
                names = tuple(mapping)
                if baseline is None:
                    baseline = names
                elif names != baseline:
                    raise ValidationError(f"Audience list/order differs in {self.paths[region].name}.")
                for audience, sheet_name in mapping.items():
                    ws = self.workbook(region)[sheet_name]
                    if not ws["B8"].value or not ws["B9"].value:
                        raise ValidationError(f"Missing B8 or B9 in {self.paths[region].name} / {audience}.")
                    for row in range(16, 101):
                        for col in "DIKLMNPQRSUVW":
                            if not isinstance(ws[f"{col}{row}"].value, (int, float)):
                                raise ValidationError(
                                    f"Expected numeric value at {col}{row} in {self.paths[region].name} / {audience}."
                                )
                    self.slide2_affinity(region, audience)
                    for start in (40, 46, 52):
                        self.competitor_set(region, audience, start)
            except Exception as exc:
                errors.append(str(exc))
        return errors


def footer_lines(repo: WorkbookRepository, region: str, audience: str, platform: str, all_platforms=False):
    ws = repo.sheet(region, audience)
    general = format_compact(ws["D16"].value) if repo.total_is_supported(region, audience, 16) else None
    reach_cell = f"{PLATFORM_COLUMNS[platform]['universe']}16"
    bbc = format_compact(ws[reach_cell].value) if repo.is_supported(region, audience, platform, 16) else None
    general_label = f"{audience} ({general})" if general else audience
    bbc_label = f"BBC {audience} ({bbc})" if bbc else f"BBC {audience}"
    return [
        f"Base: {general_label} // {bbc_label}",
        repo.waves(region, audience),
        repo.markets_footer(region),
        "Platform: BBC Cross Platform, Digital + TV (30 Day Reach)"
        if all_platforms
        else PLATFORM_FOOTERS[platform],
    ]
