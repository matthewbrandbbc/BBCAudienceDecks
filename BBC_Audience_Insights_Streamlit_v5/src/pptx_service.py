from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
import re
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile

from lxml import etree
from openpyxl import load_workbook

from .config import OUTPUT_DIR, TEMPLATE_PATH
from .presentation_model import PresentationModel

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}
P = f"{{{NS['p']}}}"
A = f"{{{NS['a']}}}"
C = f"{{{NS['c']}}}"


class ExportValidationError(RuntimeError):
    pass


def safe_filename(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")


def _replace_text_nodes(root, replacements: dict[str, str]):
    ordered = sorted(replacements.items(), key=lambda pair: len(pair[0]), reverse=True)
    # Long template sentences can be split across several PowerPoint runs.
    # Resolve them at shape level first, preserving the first run's formatting.
    long_replacements = [
        (marker, value)
        for marker, value in ordered
        if len(marker) >= 40 and not marker.startswith("__")
    ]
    for shape in root.xpath(".//p:sp", namespaces=NS):
        nodes = shape.xpath(".//a:t", namespaces=NS)
        combined = "".join((node.text or "") for node in nodes)
        updated = combined
        for marker, value in long_replacements:
            updated = updated.replace(marker, value)
        if updated != combined and nodes:
            nodes[0].text = updated
            for node in nodes[1:]:
                node.text = ""
    for node in root.xpath(".//a:t", namespaces=NS):
        text = node.text or ""
        for marker, value in ordered:
            # The template often includes a literal M after volume markers.
            text = text.replace(f"{marker}M", value)
            text = text.replace(marker, value)
        node.text = text
    # Some template suffixes are held in a separate run, e.g. "(N16)" + "M".
    # Once the replacement already contains M/K, remove the orphan suffix.
    for shape in root.xpath(".//p:sp", namespaces=NS):
        nodes = shape.xpath(".//a:t", namespaces=NS)
        for idx in range(1, len(nodes)):
            previous = (nodes[idx - 1].text or "").rstrip()
            current = (nodes[idx].text or "").strip()
            if current in {"M", "K"} and previous.endswith(current):
                nodes[idx].text = ""


def _contains_suppressed_marker(text: str, markers: set[str]) -> bool:
    return any(marker in text for marker in markers)


def _remove_suppressed_items(root, markers: set[str], slide_number: int):
    """Remove complete metric rows/blocks before placeholder replacement."""
    if not markers:
        return

    # Remove complete rows from native PowerPoint tables and shrink the table
    # frame by the deleted row heights so no blank band remains.
    for table in root.xpath(".//a:tbl", namespaces=NS):
        removed_height = 0
        for row in list(table.findall(f"./{A}tr")):
            text = "".join(row.xpath(".//a:t/text()", namespaces=NS))
            if _contains_suppressed_marker(text, markers):
                removed_height += int(row.get("h", "0"))
                table.remove(row)
        if removed_height:
            ancestor = table.getparent()
            while ancestor is not None and ancestor.tag != f"{P}graphicFrame":
                ancestor = ancestor.getparent()
            remaining_rows = table.findall(f"./{A}tr")
            if ancestor is not None and len(remaining_rows) <= 1:
                if ancestor.getparent() is not None:
                    ancestor.getparent().remove(ancestor)
            elif ancestor is not None:
                ext = ancestor.find(f"./{P}xfrm/{A}ext")
                if ext is not None:
                    ext.set("cy", str(max(1, int(ext.get("cy", "1")) - removed_height)))

    if slide_number == 2:
        labels = []
        if "(N16)" in markers:
            labels.append("Monthly Digital Reach")
        if "(S16)" in markers:
            labels.append("Monthly TV Reach")
        for shape in list(root.xpath(".//p:sp", namespaces=NS)):
            text = "".join(shape.xpath(".//a:t/text()", namespaces=NS)).strip()
            if text in labels and shape.getparent() is not None:
                shape.getparent().remove(shape)

    # Slide 5 holds two statistics in each text box. Remove the number, label
    # and spacing belonging to a suppressed metric while preserving its valid
    # neighbour.
    if slide_number == 5:
        for shape in list(root.xpath(".//p:sp", namespaces=NS)):
            body = shape.find(f"./{P}txBody")
            if body is None:
                continue
            paragraphs = list(body.findall(f"./{A}p"))
            marker_indexes = [
                idx
                for idx, para in enumerate(paragraphs)
                if "(" in "".join(para.xpath(".//a:t/text()", namespaces=NS))
            ]
            remove = set()
            for pos, start in enumerate(marker_indexes):
                marker_text = "".join(paragraphs[start].xpath(".//a:t/text()", namespaces=NS))
                if not _contains_suppressed_marker(marker_text, markers):
                    continue
                end = marker_indexes[pos + 1] if pos + 1 < len(marker_indexes) else len(paragraphs)
                remove.update(range(start, end))
            for idx in sorted(remove, reverse=True):
                body.remove(paragraphs[idx])
            visible = "".join(shape.xpath(".//a:t/text()", namespaces=NS)).strip()
            if not visible and shape.getparent() is not None:
                shape.getparent().remove(shape)

    # Slide 9's AI statements are self-contained grouped cards. Remove the
    # complete card so no orphan label remains after a statistic is suppressed.
    if slide_number == 9:
        for group in list(root.xpath(".//p:grpSp", namespaces=NS)):
            text = "".join(group.xpath(".//a:t/text()", namespaces=NS))
            if _contains_suppressed_marker(text, markers) and group.getparent() is not None:
                group.getparent().remove(group)

    # Other non-table statistics are standalone text boxes (e.g. slide 2's
    # headline reach figures). Remove the complete box rather than leaving a
    # label, suffix or placeholder behind.
    if slide_number not in {5, 9}:
        for shape in list(root.xpath(".//p:sp", namespaces=NS)):
            text = "".join(shape.xpath(".//a:t/text()", namespaces=NS))
            if _contains_suppressed_marker(text, markers) and shape.getparent() is not None:
                shape.getparent().remove(shape)


def _remove_duplicate_suppression_messages(root):
    """Keep one low-base notice when a template repeats the same title box."""
    seen = set()
    for shape in list(root.xpath(".//p:sp", namespaces=NS)):
        text = "".join(shape.xpath(".//a:t/text()", namespaces=NS)).strip()
        if not text.startswith("No ") or "statistics are shown" not in text:
            continue
        if text in seen and shape.getparent() is not None:
            shape.getparent().remove(shape)
        else:
            seen.add(text)


def _enable_text_wrap(root):
    """Wrap and shrink long text within its existing template text box."""
    for shape in root.xpath(".//p:sp", namespaces=NS):
        name = shape.xpath("./p:nvSpPr/p:cNvPr/@name", namespaces=NS)
        # Generated footers have a dedicated, bottom-anchored fit rule. Do not
        # overwrite it with the generic long-copy rule below.
        if name and name[0].startswith("Generated Footer"):
            continue
        text = "".join(shape.xpath(".//a:t/text()", namespaces=NS)).strip()
        body = shape.find(f"./{P}txBody")
        if body is None:
            continue
        body_pr = body.find(f"./{A}bodyPr")
        if body_pr is None:
            continue
        body_pr.set("wrap", "square")
        if len(text) < 48:
            continue
        for child in list(body_pr):
            if child.tag in {f"{A}noAutofit", f"{A}spAutoFit", f"{A}normAutofit"}:
                body_pr.remove(child)
        autofit = etree.SubElement(body_pr, f"{A}normAutofit")
        autofit.set("fontScale", "75000" if len(text) >= 100 else "85000")
        autofit.set("lnSpcReduction", "15000" if len(text) >= 100 else "10000")


def _add_audience_bubble(slide_root, label: str):
    sp_tree = slide_root.find(f".//{P}spTree")
    source = None
    for shape in slide_root.xpath(".//p:sp", namespaces=NS):
        if "".join(shape.xpath(".//a:t/text()", namespaces=NS)).strip() == "Reach by Region":
            source = deepcopy(shape)
            break
    if source is None:
        raise ExportValidationError("Could not locate the Reach by Region title bubble.")
    ids = [int(v) for v in slide_root.xpath(".//p:cNvPr/@id", namespaces=NS) if v.isdigit()]
    c_nv = source.find(f"./{P}nvSpPr/{P}cNvPr")
    c_nv.set("id", str(max(ids or [1]) + 1))
    c_nv.set("name", "Selected Audience Bubble")
    xfrm = source.find(f"./{P}spPr/{A}xfrm")
    xfrm.find(f"./{A}off").set("x", "1635000")
    xfrm.find(f"./{A}ext").set("cx", "2200000")
    nodes = source.xpath(".//a:t", namespaces=NS)
    if nodes:
        nodes[0].text = label
        for node in nodes[1:]:
            node.text = ""
    sp_tree.append(source)


def _set_fonts(root):
    for node in root.xpath(".//*[@typeface]", namespaces=NS):
        face = node.get("typeface", "")
        if "Reith Serif" in face:
            node.set("typeface", "Georgia")
        elif "Reith" in face:
            node.set("typeface", "Arial")


def _footer_shape(template_root, lines: list[str], shape_id: int, slide_number: int):
    source = None
    for shape in template_root.xpath(".//p:sp", namespaces=NS):
        name = shape.xpath("./p:nvSpPr/p:cNvPr/@name", namespaces=NS)
        text = "".join(shape.xpath(".//a:t/text()", namespaces=NS))
        if name and name[0] == "TextBox 64" and "Interviews" in text:
            source = deepcopy(shape)
            break
    if source is None:
        raise ExportValidationError("Could not locate the template footer text box.")
    c_nv = source.find(f"./{P}nvSpPr/{P}cNvPr")
    c_nv.set("id", str(shape_id))
    c_nv.set("name", f"Generated Footer {shape_id}")
    xfrm = source.find(f"./{P}spPr/{A}xfrm")
    if slide_number == 2:
        x, y, cx, cy, font_size = "9020000", "5750000", "2850000", "980000", "500"
    else:
        x, y, cx, cy, font_size = "334963", "6330000", "11512588", "470000", "475"
    xfrm.find(f"./{A}off").set("x", x)
    xfrm.find(f"./{A}off").set("y", y)
    xfrm.find(f"./{A}ext").set("cx", cx)
    xfrm.find(f"./{A}ext").set("cy", cy)
    tx_body = source.find(f"./{P}txBody")
    body_pr = tx_body.find(f"./{A}bodyPr")
    if body_pr is None:
        body_pr = etree.Element(f"{A}bodyPr")
        tx_body.insert(0, body_pr)
    body_pr.set("anchor", "b")
    body_pr.set("wrap", "square")
    body_pr.set("lIns", "0")
    body_pr.set("rIns", "0")
    body_pr.set("tIns", "0")
    body_pr.set("bIns", "0")
    for child in list(body_pr):
        if child.tag in {f"{A}noAutofit", f"{A}spAutoFit", f"{A}normAutofit"}:
            body_pr.remove(child)
    footer_fit = etree.SubElement(body_pr, f"{A}normAutofit")
    footer_fit.set("fontScale", "95000")
    footer_fit.set("lnSpcReduction", "12000")
    for child in list(tx_body):
        if child.tag == f"{A}p":
            tx_body.remove(child)
    for line in lines:
        para = etree.SubElement(tx_body, f"{A}p")
        ppr = etree.SubElement(para, f"{A}pPr")
        ppr.set("algn", "l")
        ppr.set("marL", "0")
        ppr.set("marR", "0")
        ppr.set("indent", "0")
        line_spacing = etree.SubElement(ppr, f"{A}lnSpc")
        etree.SubElement(line_spacing, f"{A}spcPct").set("val", "90000")
        before = etree.SubElement(ppr, f"{A}spcBef")
        etree.SubElement(before, f"{A}spcPts").set("val", "0")
        after = etree.SubElement(ppr, f"{A}spcAft")
        etree.SubElement(after, f"{A}spcPts").set("val", "0")
        run = etree.SubElement(para, f"{A}r")
        rpr = etree.SubElement(run, f"{A}rPr")
        rpr.set("lang", "en-GB")
        rpr.set("sz", font_size)
        latin = etree.SubElement(rpr, f"{A}latin")
        latin.set("typeface", "Arial")
        text = etree.SubElement(run, f"{A}t")
        text.text = line
        etree.SubElement(para, f"{A}endParaRPr").set("lang", "en-GB")
    return source


def _apply_footer(slide_root, footer_template_root, lines: list[str], slide_number: int):
    sp_tree = slide_root.find(f".//{P}spTree")
    ids = [int(v) for v in slide_root.xpath(".//p:cNvPr/@id", namespaces=NS) if v.isdigit()]
    generated = _footer_shape(footer_template_root, lines, max(ids or [1]) + 1, slide_number)
    # Remove the old marketing footer on slides 3/4.
    for shape in list(sp_tree):
        text = "".join(shape.xpath(".//a:t/text()", namespaces=NS)) if hasattr(shape, "xpath") else ""
        if "1750 Interviews" in text:
            sp_tree.remove(shape)
    sp_tree.append(generated)


def _update_rank_statements(slide_root, charts):
    for shape in slide_root.xpath(".//p:sp", namespaces=NS):
        text = "".join(shape.xpath(".//a:t/text()", namespaces=NS))
        if "ranks #2" not in text:
            continue
        off = shape.find(f"./{P}spPr/{A}xfrm/{A}off")
        x = int(off.get("x", "0")) if off is not None else 0
        platform = "Cross Platform" if x < 4_000_000 else "Digital" if x < 8_000_000 else "TV"
        if charts[platform]["rank"] is None:
            if shape.getparent() is not None:
                shape.getparent().remove(shape)
            continue
        context = {
            "Cross Platform": "30-day Digital + TV Reach",
            "Digital": "30-day Digital Reach",
            "TV": "30-day TV Reach",
        }[platform]
        new_text = f"BBC {platform} audience ranks {charts[platform]['ordinal']} amongst the competitive set. ({context})"
        nodes = shape.xpath(".//a:t", namespaces=NS)
        if nodes:
            nodes[0].text = new_text
            for node in nodes[1:]:
                node.text = ""


def _remove_empty_charts(slide_root, charts):
    for frame in list(slide_root.xpath(".//p:graphicFrame[.//c:chart]", namespaces=NS)):
        off = frame.find(f"./{P}xfrm/{A}off")
        x = int(off.get("x", "0")) if off is not None else 0
        platform = "Cross Platform" if x < 4_000_000 else "Digital" if x < 8_000_000 else "TV"
        if not charts[platform]["items"] and frame.getparent() is not None:
            frame.getparent().remove(frame)
    empty_titles = {
        f"{platform} Reach" for platform, chart in charts.items() if not chart["items"]
    }
    for shape in list(slide_root.xpath(".//p:sp", namespaces=NS)):
        text = "".join(shape.xpath(".//a:t/text()", namespaces=NS)).strip()
        if text in empty_titles and shape.getparent() is not None:
            shape.getparent().remove(shape)


def _set_solid_fill(dpt, red: bool):
    sppr = dpt.find(f"./{C}spPr")
    if sppr is None:
        sppr = etree.SubElement(dpt, f"{C}spPr")
    old = sppr.find(f"./{A}solidFill")
    if old is not None:
        sppr.remove(old)
    fill = etree.Element(f"{A}solidFill")
    color = etree.SubElement(fill, f"{A}srgbClr" if red else f"{A}schemeClr")
    color.set("val", "C00000" if red else "tx1")
    sppr.insert(0, fill)


def _replace_cache(cache, values, numeric=False):
    for pt in list(cache.findall(f"./{C}pt")):
        cache.remove(pt)
    count = cache.find(f"./{C}ptCount")
    if count is None:
        count = etree.SubElement(cache, f"{C}ptCount")
    count.set("val", str(len(values)))
    for idx, value in enumerate(values):
        pt = etree.SubElement(cache, f"{C}pt")
        pt.set("idx", str(idx))
        etree.SubElement(pt, f"{C}v").text = str(float(value) if numeric else value)


def _update_chart_xml(xml_bytes: bytes, chart: dict) -> bytes:
    root = etree.fromstring(xml_bytes)
    items = chart["items"]
    cat_cache = root.find(".//c:ser/c:cat/c:strRef/c:strCache", namespaces=NS)
    num_cache = root.find(".//c:ser/c:val/c:numRef/c:numCache", namespaces=NS)
    _replace_cache(cat_cache, [item["name"] for item in items])
    _replace_cache(num_cache, [item["value"] for item in items], numeric=True)
    labels = root.findall(".//c:ser/c:dLbls/c:dLbl", namespaces=NS)
    for idx, label in enumerate(labels):
        if idx >= len(items):
            label.getparent().remove(label)
            continue
        text_nodes = label.xpath(".//a:t", namespaces=NS)
        if text_nodes:
            text_nodes[0].text = items[idx]["label"]
    for dpt in root.findall(".//c:ser/c:dPt", namespaces=NS):
        idx_node = dpt.find(f"./{C}idx")
        idx = int(idx_node.get("val", "-1")) if idx_node is not None else -1
        if idx >= len(items):
            dpt.getparent().remove(dpt)
            continue
        _set_solid_fill(dpt, 0 <= idx < len(items) and items[idx]["name"].casefold() == "bbc")
    _set_fonts(root)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _update_embedded_workbook(data: bytes, chart: dict) -> bytes:
    source = BytesIO(data)
    wb = load_workbook(source)
    ws = wb["Sheet1"]
    ws["A1"] = "Organisation"
    ws["B1"] = "Reach"
    for row in range(2, 8):
        ws.cell(row, 1).value = None
        ws.cell(row, 2).value = None
    for row, item in enumerate(chart["items"], 2):
        ws.cell(row, 1).value = item["name"]
        ws.cell(row, 2).value = item["value"]
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pptx(model: PresentationModel, output_path: Path | None = None) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if output_path is None:
        output_path = OUTPUT_DIR / (
            f"BBC_Audience_Deck_{safe_filename(model.audience)}_"
            f"{safe_filename(model.region)}_{safe_filename(model.platform)}.pptx"
        )
    with ZipFile(TEMPLATE_PATH, "r") as zin:
        package = {name: zin.read(name) for name in zin.namelist()}
    footer_template = etree.fromstring(package["ppt/slides/slide3.xml"])
    for slide in model.slides:
        name = f"ppt/slides/slide{slide.number}.xml"
        root = etree.fromstring(package[name])
        _remove_suppressed_items(root, slide.suppressed_markers, slide.number)
        _replace_text_nodes(root, slide.replacements)
        _remove_duplicate_suppression_messages(root)
        if slide.number == 3:
            _add_audience_bubble(root, slide.replacements["__AUDIENCE_BUBBLE__"])
        if slide.number == 4:
            _update_rank_statements(root, slide.charts)
            _remove_empty_charts(root, slide.charts)
        if slide.footer:
            _apply_footer(root, footer_template, slide.footer, slide.number)
        _enable_text_wrap(root)
        _set_fonts(root)
        package[name] = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

    chart_names = {
        # slide4.xml places chart2/rId4 on the left and chart1/rId3 in the middle.
        "Cross Platform": ("ppt/charts/chart2.xml", "ppt/embeddings/Microsoft_Excel_Worksheet1.xlsx"),
        "Digital": ("ppt/charts/chart1.xml", "ppt/embeddings/Microsoft_Excel_Worksheet.xlsx"),
        "TV": ("ppt/charts/chart3.xml", "ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx"),
    }
    slide4 = model.slides[3]
    for platform, (chart_path, workbook_path) in chart_names.items():
        chart = slide4.charts[platform]
        package[chart_path] = _update_chart_xml(package[chart_path], chart)
        package[workbook_path] = _update_embedded_workbook(package[workbook_path], chart)

    for name, data in list(package.items()):
        if name.endswith(".xml"):
            data = data.replace(b"BBC Reith Serif", b"Georgia")
            data = data.replace(b"BBC Reith Sans", b"Arial")
            data = data.replace(b"&lt;#&gt;", b"")
            data = data.replace("‹#›".encode("utf-8"), b"")
            data = data.replace(
                b"application/vnd.openxmlformats-officedocument.presentationml.template.main+xml",
                b"application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml",
            )
            package[name] = data

    with ZipFile(output_path, "w", ZIP_DEFLATED) as zout:
        for name, data in package.items():
            zout.writestr(name, data)
    validate_pptx(output_path)
    return output_path


def validate_pptx(path: Path):
    marker = re.compile(r"\([A-W]{1,2}\d{1,3}\)")
    with ZipFile(path) as z:
        slides = [n for n in z.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", n)]
        if len(slides) != 9:
            raise ExportValidationError(f"Generated PPTX contains {len(slides)} slides instead of 9.")
        unresolved = []
        for name in slides:
            root = etree.fromstring(z.read(name))
            text = " ".join(root.xpath(".//a:t/text()", namespaces=NS))
            unresolved.extend(marker.findall(text))
        if unresolved:
            raise ExportValidationError(f"Unresolved placeholders remain: {sorted(set(unresolved))}")
        embeddings = [n for n in z.namelist() if n.startswith("ppt/embeddings/") and n.endswith(".xlsx")]
        expected = {
            "ppt/embeddings/Microsoft_Excel_Worksheet.xlsx",
            "ppt/embeddings/Microsoft_Excel_Worksheet1.xlsx",
            "ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx",
        }
        if set(embeddings) != expected:
            raise ExportValidationError("PPTX contains unexpected embedded workbooks.")
        if b"template.main+xml" in z.read("[Content_Types].xml"):
            raise ExportValidationError("Output still declares itself as a PowerPoint template.")
